"""独立的 XLS 转 XLSX 数据清理分列模块。

本模块不导入 processor.py 或 excel_clean_split.py。旧版 XLS 的格式保真转换与
TextToColumns 均交给 Microsoft Excel COM；openpyxl 仅用于 data_only=False 的
结构读取、原始值记录和最终结果校验。
"""
from __future__ import annotations

import os
from pathlib import Path
import queue
import subprocess
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import winreg

from openpyxl import load_workbook


REQUIRED_HEADERS = ("店铺名称", "昨日交易额", "昨日订单数")


class XlsCleanSplitError(RuntimeError):
    """可安全显示给用户的转换或校验错误。"""


def desktop_path() -> Path:
    """读取 Windows 当前用户配置的桌面目录。"""
    try:
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
        ) as key:
            value, _ = winreg.QueryValueEx(key, "Desktop")
        return Path(os.path.expandvars(value)).expanduser()
    except OSError as exc:
        raise XlsCleanSplitError("无法获取 Windows 桌面路径。") from exc


def _normalized_header(value) -> str | None:
    return value.strip() if isinstance(value, str) else None


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    """允许标题前后空格，但规范化后必须与三个字段精确相等。"""
    for row_index in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            header = _normalized_header(sheet.cell(row_index, column_index).value)
            if header in REQUIRED_HEADERS:
                if header in found:
                    raise XlsCleanSplitError(f"标题行存在重复字段：{header}")
                found[header] = column_index
        if all(header in found for header in REQUIRED_HEADERS):
            return row_index, found
    raise XlsCleanSplitError("未找到店铺名称、昨日交易额、昨日订单数三个标题。")


def inspect_converted_workbook(path: Path) -> tuple[int, int, int, dict[str, int], dict[str, list[int]]]:
    """用 openpyxl(data_only=False) 读取并记录所有原始单元格值。"""
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook.active
        header_row, columns = find_header_row(sheet)
        # 明确读取所有原始单元格，而非只读三列。
        original_values = {
            cell.coordinate: cell.value
            for row in sheet.iter_rows()
            for cell in row
        }
        nonempty_rows: dict[str, list[int]] = {}
        for header in ("昨日交易额", "昨日订单数"):
            column_index = columns[header]
            nonempty_rows[header] = [
                row_index
                for row_index in range(header_row + 1, sheet.max_row + 1)
                if original_values.get(sheet.cell(row_index, column_index).coordinate) is not None
            ]
        return workbook.index(sheet) + 1, header_row, sheet.max_column, columns, nonempty_rows
    finally:
        workbook.close()


def verify_output(path: Path) -> None:
    """校验最终三列、标题顺序，并确保结果仍可由 openpyxl 读取。"""
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        sheet = workbook.active
        header_row, columns = find_header_row(sheet)
        expected = {header: index for index, header in enumerate(REQUIRED_HEADERS, start=1)}
        if columns != expected or sheet.max_column != 3:
            raise XlsCleanSplitError("输出列结构校验失败，未严格生成指定的 A、B、C 三列。")
        # 再次逐个读取，确保保存后的内容结构完整可访问。
        for row in sheet.iter_rows(min_row=header_row):
            for cell in row:
                _ = cell.value
    finally:
        workbook.close()


def _ps_quote(value: Path | str) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _run_excel_script(script: str) -> None:
    """运行独立 PowerShell/Excel COM 脚本，不依赖 pywin32。"""
    script_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".ps1", delete=False) as handle:
            script_path = Path(handle.name)
        script_path.write_text(script, encoding="utf-8-sig")
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass",
             "-File", str(script_path)],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0), timeout=300,
        )
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout).strip()
            raise XlsCleanSplitError(
                "Microsoft Excel 转换失败。请确认电脑已安装 Excel，且输入/输出文件未被占用。"
                + (f"\n{detail}" if detail else "")
            )
    except FileNotFoundError as exc:
        raise XlsCleanSplitError("无法启动 Windows PowerShell。") from exc
    except subprocess.TimeoutExpired as exc:
        raise XlsCleanSplitError("Excel 转换超过 5 分钟，操作已停止。") from exc
    finally:
        if script_path:
            script_path.unlink(missing_ok=True)


def _conversion_script(input_path: Path, output_path: Path) -> str:
    """由 Excel 原生 SaveAs 将 BIFF8 XLS 转成 Open XML XLSX。"""
    return f"""
$ErrorActionPreference = 'Stop'
$excel = $null
$book = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open({_ps_quote(input_path)}, 0, $true)
    $book.SaveAs({_ps_quote(output_path)}, 51)
    $book.Close($false)
    $book = $null
}} finally {{
    if ($null -ne $book) {{ $book.Close($false) }}
    if ($null -ne $excel) {{ $excel.Quit() }}
    if ($null -ne $book) {{ [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) }}
    if ($null -ne $excel) {{ [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) }}
    [GC]::Collect(); [GC]::WaitForPendingFinalizers()
}}
"""


def _processing_script(path: Path, sheet_index: int, header_row: int, max_column: int,
                       nonempty_rows: dict[str, list[int]]) -> str:
    sales_rows = ",".join(str(row) for row in nonempty_rows["昨日交易额"])
    order_rows = ",".join(str(row) for row in nonempty_rows["昨日订单数"])
    return f"""
$ErrorActionPreference = 'Stop'
$excel = $null
$book = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open({_ps_quote(path)}, 0, $false)
    $sheet = $book.Worksheets.Item({sheet_index})
    $wanted = @('店铺名称','昨日交易额','昨日订单数')
    for ($column = {max_column}; $column -ge 1; $column--) {{
        $header = ([string]$sheet.Cells.Item({header_row}, $column).Value2).Trim()
        if ($wanted -notcontains $header) {{ $sheet.Columns.Item($column).Delete() }}
    }}
    for ($target = 1; $target -le 3; $target++) {{
        $current = 0
        for ($column = 1; $column -le 3; $column++) {{
            if (([string]$sheet.Cells.Item({header_row}, $column).Value2).Trim() -eq $wanted[$target - 1]) {{
                $current = $column; break
            }}
        }}
        if ($current -eq 0) {{ throw '保留字段在转换过程中丢失。' }}
        if ($current -ne $target) {{
            $sheet.Columns.Item($current).Cut()
            $sheet.Columns.Item($target).Insert(-4161)
            $excel.CutCopyMode = 0
        }}
    }}
    $salesRows = @({sales_rows})
    foreach ($row in $salesRows) {{
        $cell = $sheet.Cells.Item($row, 2)
        $cell.TextToColumns($cell, 1, 1, $false, $false, $false, $false, $false, $false, '', @(@(1,1)), '.', ',', $true)
    }}
    $orderRows = @({order_rows})
    foreach ($row in $orderRows) {{
        $cell = $sheet.Cells.Item($row, 3)
        $cell.TextToColumns($cell, 1, 1, $false, $false, $false, $false, $false, $false, '', @(@(1,1)), '.', ',', $true)
    }}
    $book.Save()
    $book.Close($true)
    $book = $null
}} finally {{
    if ($null -ne $book) {{ $book.Close($false) }}
    if ($null -ne $excel) {{ $excel.Quit() }}
    if ($null -ne $book) {{ [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($book) }}
    if ($null -ne $excel) {{ [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) }}
    [GC]::Collect(); [GC]::WaitForPendingFinalizers()
}}
"""


def process_xls(input_path: Path, output_path: Path) -> Path:
    """转换、记录原值、原生分列、校验并原子输出。"""
    if input_path.suffix.lower() != ".xls":
        raise XlsCleanSplitError("请选择扩展名为 .xls 的旧版 Excel 文件。")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=output_path.parent, prefix="xls_clean_split_") as folder:
        working = Path(folder) / (input_path.stem + ".xlsx")
        _run_excel_script(_conversion_script(input_path.resolve(), working.resolve()))
        sheet_index, header_row, max_column, _columns, nonempty_rows = inspect_converted_workbook(working)
        _run_excel_script(
            _processing_script(working.resolve(), sheet_index, header_row, max_column, nonempty_rows)
        )
        verify_output(working)
        working.replace(output_path)
    return output_path


class XlsCleanSplitWindow(tk.Toplevel):
    """XLS 转换清理功能的独立窗口。"""

    def __init__(self, master):
        super().__init__(master)
        self.title("XLS转XLSX数据清理分列助手")
        self.geometry("680x320")
        self.minsize(580, 280)
        self.input_path: Path | None = None
        self.events: queue.Queue[tuple[str, str]] = queue.Queue()
        self._build_ui()
        self.after(100, self._drain_events)

    def _build_ui(self) -> None:
        frame = ttk.Frame(self, padding=18)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="XLS转XLSX数据清理分列助手",
                  font=("Microsoft YaHei UI", 16, "bold")).pack(anchor="w")
        ttk.Label(frame, text="使用 Microsoft Excel 原生转换和分列，结果自动保存到桌面。",
                  wraplength=620).pack(anchor="w", pady=(4, 16))
        ttk.Button(frame, text="1. 选择XLS文件", command=self.select_file).pack(anchor="w")
        self.file_label = ttk.Label(frame, text="尚未选择", foreground="#555")
        self.file_label.pack(anchor="w", pady=(6, 14))
        self.start_button = ttk.Button(frame, text="2. 开始转换并处理", command=self.start)
        self.start_button.pack(anchor="w")
        self.progress = ttk.Progressbar(frame, mode="indeterminate")
        self.progress.pack(fill="x", pady=(14, 8))
        self.status_label = ttk.Label(frame, text="运行要求：电脑已安装 Microsoft Excel")
        self.status_label.pack(anchor="w")

    def select_file(self) -> None:
        path = filedialog.askopenfilename(parent=self, title="选择XLS文件",
                                          filetypes=[("Excel 97-2003 工作簿", "*.xls")])
        if path:
            self.input_path = Path(path)
            self.file_label.config(text=str(self.input_path))

    def start(self) -> None:
        if not self.input_path:
            messagebox.showwarning("缺少文件", "请先选择一个 .xls 文件。", parent=self)
            return
        self.start_button.config(state="disabled")
        self.progress.start(10)
        self.status_label.config(text="正在转换和处理，请勿打开相关文件……")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        try:
            output = desktop_path() / f"{self.input_path.stem}.xlsx"
            process_xls(self.input_path, output)
            self.events.put(("done", str(output)))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def _drain_events(self) -> None:
        try:
            while True:
                kind, text = self.events.get_nowait()
                self.progress.stop()
                self.start_button.config(state="normal")
                if kind == "done":
                    self.status_label.config(text=f"处理完成：{text}")
                    messagebox.showinfo("完成", f"文件已输出到：\n{text}", parent=self)
                else:
                    self.status_label.config(text=f"处理失败：{text}")
                    messagebox.showerror("处理失败", text, parent=self)
        except queue.Empty:
            pass
        if self.winfo_exists():
            self.after(100, self._drain_events)
