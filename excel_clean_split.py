"""独立的 Excel 数据清理分列功能；不依赖日报处理模块。"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from copy import copy
from pathlib import Path
import os
import queue
import re
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import winreg

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REQUIRED_HEADERS = ("店铺名称", "昨日交易额", "昨日订单数")
SPLIT_HEADERS = {"昨日交易额", "昨日订单数"}
NUMBER_TEXT = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?$")


class CleanSplitError(RuntimeError):
    """可直接显示给用户的清理错误。"""


def desktop_path() -> Path:
    """从 Windows 注册表获取当前用户的真实桌面目录。"""
    try:
        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
        ) as key:
            value, _ = winreg.QueryValueEx(key, "Desktop")
        return Path(os.path.expandvars(value)).expanduser()
    except OSError as exc:
        raise CleanSplitError("无法获取 Windows 桌面路径。") from exc


def find_header_row(sheet) -> tuple[int, dict[str, int]]:
    """精确查找同时包含三个指定标题的标题行。"""
    for row_index in range(1, sheet.max_row + 1):
        found: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            value = sheet.cell(row_index, column_index).value
            if isinstance(value, str) and value in REQUIRED_HEADERS:
                found[value] = column_index
        if all(header in found for header in REQUIRED_HEADERS):
            return row_index, found
    raise CleanSplitError("未找到同时包含“店铺名称、昨日交易额、昨日订单数”的标题行。")


def excel_general_value(value):
    """模拟“分列-常规”对单个非空文本数字的类型转换。"""
    if not isinstance(value, str) or value == "" or value.startswith("="):
        return value
    candidate = value.strip()
    if not candidate or not NUMBER_TEXT.fullmatch(candidate):
        return value
    try:
        number = Decimal(candidate)
    except InvalidOperation:
        return value
    if not number.is_finite():
        return value
    return int(number) if number == number.to_integral_value() else float(number)


def clean_split_workbook(input_path: Path, output_path: Path) -> Path:
    """仅清理选定工作簿；不会调用或修改任何日报处理代码。"""
    workbook = load_workbook(input_path, data_only=False)
    try:
        sheet = workbook.active
        header_row, columns = find_header_row(sheet)
        keep_columns = set(columns.values())
        ordered_kept = sorted(keep_columns)
        kept_dimensions = [
            copy(sheet.column_dimensions[get_column_letter(column_index)])
            for column_index in ordered_kept
        ]
        for column_index in range(sheet.max_column, 0, -1):
            if column_index not in keep_columns:
                sheet.delete_cols(column_index, 1)

        # delete_cols 不迁移 column_dimensions，需按保留列的新位置恢复。
        sheet.column_dimensions.clear()
        for new_index, dimension in enumerate(kept_dimensions, start=1):
            letter = get_column_letter(new_index)
            dimension.index = letter
            dimension.min = new_index
            dimension.max = new_index
            sheet.column_dimensions[letter] = dimension

        _, remaining = find_header_row(sheet)
        for header in SPLIT_HEADERS:
            column_index = remaining[header]
            for row_index in range(header_row + 1, sheet.max_row + 1):
                cell = sheet.cell(row_index, column_index)
                original = cell.value
                if original is None or original == "":
                    continue
                converted = excel_general_value(original)
                if converted != original or type(converted) is not type(original):
                    cell.value = converted

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            dir=output_path.parent, prefix=".~clean_split_", suffix=".xlsx", delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
        try:
            workbook.save(temporary_path)
            temporary_path.replace(output_path)
        finally:
            temporary_path.unlink(missing_ok=True)
    finally:
        workbook.close()
    return output_path


class CleanSplitWindow(tk.Toplevel):
    """与日报页面隔离的独立清理窗口。"""

    def __init__(self, master):
        super().__init__(master)
        self.title("Excel数据清理分列助手")
        self.geometry("650x300")
        self.minsize(560, 260)
        self.input_path: Path | None = None
        self.events: queue.Queue[tuple[str, str]] = queue.Queue()
        self._build_ui()
        self.after(100, self._drain_events)

    def _build_ui(self) -> None:
        frame = ttk.Frame(self, padding=18)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Excel数据清理分列助手", font=("Microsoft YaHei UI", 16, "bold")).pack(anchor="w")
        ttk.Label(frame, text="仅保留三个指定字段，并将交易额、订单数执行常规分列转换；空值不写入。",
                  wraplength=600).pack(anchor="w", pady=(4, 16))
        ttk.Button(frame, text="1. 选择输入Excel文件", command=self.select_file).pack(anchor="w")
        self.file_label = ttk.Label(frame, text="尚未选择", foreground="#555")
        self.file_label.pack(anchor="w", pady=(6, 14))
        self.start_button = ttk.Button(frame, text="2. 开始处理", command=self.start)
        self.start_button.pack(anchor="w")
        self.progress = ttk.Progressbar(frame, mode="indeterminate")
        self.progress.pack(fill="x", pady=(14, 8))
        self.status_label = ttk.Label(frame, text="输出位置：Windows桌面")
        self.status_label.pack(anchor="w")

    def select_file(self) -> None:
        path = filedialog.askopenfilename(parent=self, title="选择Excel文件",
                                          filetypes=[("Excel 工作簿", "*.xlsx")])
        if path:
            self.input_path = Path(path)
            self.file_label.config(text=str(self.input_path))

    def start(self) -> None:
        if not self.input_path:
            messagebox.showwarning("缺少文件", "请先选择一个 XLSX 文件。", parent=self)
            return
        self.start_button.config(state="disabled")
        self.progress.start(10)
        self.status_label.config(text="正在处理，请勿打开目标文件……")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self) -> None:
        try:
            output = desktop_path() / self.input_path.name
            clean_split_workbook(self.input_path, output)
            self.events.put(("done", str(output)))
        except Exception as exc:
            self.events.put(("error", str(exc)))

    def _drain_events(self) -> None:
        try:
            while True:
                kind, text = self.events.get_nowait()
                self.progress.stop()
                self.start_button.config(state="normal")
                if kind == "done":
                    self.status_label.config(text=f"处理完成：{text}")
                    messagebox.showinfo("完成", f"文件已输出到：\n{text}", parent=self)
                else:
                    self.status_label.config(text=f"处理失败：{text}")
                    messagebox.showerror("处理失败", text, parent=self)
        except queue.Empty:
            pass
        if self.winfo_exists():
            self.after(100, self._drain_events)
