"""月报独立处理模块。不得导入 weekly_report。"""
from __future__ import annotations

import calendar
from collections import defaultdict
from copy import copy
from datetime import date
from pathlib import Path
import re
import tempfile
from typing import Callable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from excel_utils import (
    ReportError, daily_sheet_name, duplicate_template_sheet, find_header,
    find_optional_header_column, find_summary_row, is_summary_label, keep_only_sheets,
    last_business_column, load_template_workbook, parse_date_value, snapshot_detail_row,
    solidify_formula_sheet, source_column_layout, write_complete_detail_rows,
    xlsx_sheet_names,
)
from formula_utils import normalize_division_columns, replace_sheet_references
from file_match import date_from_filename


DAILY_ALIASES = {
    "store": ("店铺名称", "店铺", "门店", "店名", "账号", "store"),
    "sales": ("本期销售额", "今日销售额", "昨日交易额", "销售额", "销售金额", "成交金额", "sales"),
    "orders": ("订单数", "昨日订单数", "订单量", "订单", "销量", "orders"),
    "refund": ("退款金额", "退款", "refund"),
}
DATED_DAILY_ALIASES = {**DAILY_ALIASES, "date": ("日期", "数据日期", "统计日期", "日报日期")}
REPORT_ALIASES = {
    "store": ("店铺名称", "店铺", "门店", "店名", "账号", "store"),
    "sales": ("本月销售额", "销售额", "销售金额", "成交金额", "sales"),
    "orders": ("订单数", "订单量", "订单", "销量", "orders"),
    "previous": ("上月销售额", "昨日销售额", "previous"),
}
REPORT_OPTIONAL_ALIASES = {
    "refund": ("退款金额", "退款", "refund"),
    "refund_rate": ("退款率", "refund_rate"),
    "comparison": ("销售额环比", "环比"),
}
MONTH_PATTERN = re.compile(r"(?:1[0-2]|[1-9])月")


_DUPLICATE_HEADER_FIELDS = {
    "\u65e5\u671f", "\u6392\u5e8f", "\u6392\u540d", "\u4f9b\u5e94\u5546", "\u4ea7\u4e1a\u5e26",
    "\u5e97\u94fa\u540d\u79f0", "\u5e97\u94fa", "\u4eca\u65e5\u9500\u552e\u989d", "\u9500\u552e\u989d",
    "\u8ba2\u5355\u6570", "\u5ba2\u5355\u4ef7", "\u6628\u65e5\u9500\u552e\u989d",
}


def _normalized_header_value(value) -> str:
    """Normalize a header value without changing the workbook itself."""
    return re.sub(r"\s+", "", str(value or "")).casefold()


def _find_second_header(sheet, header_row: int) -> int | None:
    """Return a repeated header row when four or more fields match."""
    first_header = {
        _normalized_header_value(sheet.cell(header_row, column).value)
        for column in range(1, sheet.max_column + 1)
    }
    first_header.discard("")
    known_fields = {_normalized_header_value(value) for value in _DUPLICATE_HEADER_FIELDS}
    for row in range(header_row + 1, sheet.max_row + 1):
        row_values = {
            _normalized_header_value(sheet.cell(row, column).value)
            for column in range(1, sheet.max_column + 1)
        }
        row_values.discard("")
        if len(row_values & first_header) >= 4 or len(row_values & known_fields) >= 4:
            return row
    return None


def _clean_second_header(source, cached, header_row: int,
                         logger: Callable[[str], None], log_once: bool) -> bool:
    """Remove a repeated header and everything below it from imported copies."""
    second_header = _find_second_header(source, header_row)
    if second_header is None:
        if log_once:
            logger("\u68c0\u6d4b\u7b2c\u4e8c\u8868\u5934\uff1a\n\u5426")
            logger("\u672a\u53d1\u73b0\u91cd\u590d\u8868\u5934\u3002")
        return False

    original_last_row = source.max_row
    source.delete_rows(second_header, original_last_row - second_header + 1)
    if cached.max_row >= second_header:
        cached.delete_rows(second_header, cached.max_row - second_header + 1)
    if log_once:
        logger("\u68c0\u6d4b\u7b2c\u4e8c\u8868\u5934\uff1a\n\u662f")
        logger(f"\u7b2c\u4e8c\u8868\u5934\u4f4d\u7f6e\uff1a\n\u7b2c{second_header}\u884c")
        logger(f"\u5220\u9664\u8303\u56f4\uff1a\n{second_header}\u884c\u81f3\u6700\u540e\u4e00\u884c")
    return True


def _find_month_report_header(sheet) -> tuple[int, dict[str, int]]:
    """Locate required and optional monthly fields within the first 100 rows."""
    header_row, columns = find_header(sheet, REPORT_ALIASES)
    for key, aliases in REPORT_OPTIONAL_ALIASES.items():
        column = find_optional_header_column(sheet, header_row, aliases)
        if column is not None:
            columns[key] = column
    return header_row, columns


def _collect_month(daily_paths: list[Path], year: int, month: int,
                   logger: Callable[[str], None] = lambda _: None):
    rows = []
    refund_found = False
    valid_days = {date(year, month, day) for day in range(1, calendar.monthrange(year, month)[1] + 1)}
    sources: dict[date, tuple[Path, str]] = {}
    consolidated: dict[date, list[dict]] = defaultdict(list)
    data_columns = None
    column_layout = None
    cleaning_reported: set[tuple[str, str]] = set()
    for daily_path in daily_paths:
        probe = load_workbook(daily_path, data_only=False, read_only=True)
        try:
            found_in_sheets = False
            for day in valid_days:
                name = daily_sheet_name(day)
                if name in probe.sheetnames:
                    if day in sources:
                        raise ReportError(f"日期 {day.year}/{day.month}/{day.day} 存在重复日报。")
                    sources[day] = (daily_path, name)
                    found_in_sheets = True
            file_day = date_from_filename(daily_path, year)
            if not found_in_sheets and file_day in valid_days:
                if file_day in sources:
                    raise ReportError(f"日期 {file_day.year}/{file_day.month}/{file_day.day} 存在重复日报。")
                sources[file_day] = (daily_path, probe.active.title)
        finally:
            probe.close()

        formulas = load_workbook(daily_path, data_only=False)
        cached = load_workbook(daily_path, data_only=True)
        try:
            for name in formulas.sheetnames:
                source, values = formulas[name], cached[name]
                try:
                    header_row, columns = find_header(source, DATED_DAILY_ALIASES)
                except ReportError:
                    continue
                clean_key = (str(daily_path.resolve()).casefold(), name.casefold())
                _clean_second_header(
                    source, values, header_row, logger, clean_key not in cleaning_reported
                )
                cleaning_reported.add(clean_key)
                # Cleaning changes the used range. Re-identify both the data area
                # and summary row before extracting imported daily records.
                header_row, columns = find_header(source, DATED_DAILY_ALIASES)
                find_summary_row(source, header_row)
                refund_found = refund_found or "refund" in columns
                last_column = last_business_column(source, header_row)
                if data_columns is None:
                    data_columns = dict(columns)
                    column_layout = source_column_layout(source, last_column)
                for row in range(header_row + 1, source.max_row + 1):
                    row_day = parse_date_value(source.cell(row, columns["date"]).value, year)
                    if row_day not in valid_days:
                        continue
                    store = source.cell(row, columns["store"]).value
                    if store is None or str(store).strip() == "" or is_summary_label(store):
                        continue
                    record = {"store": str(store).strip()}
                    for key in ("sales", "orders", "refund"):
                        if key not in columns:
                            continue
                        value = source.cell(row, columns[key]).value
                        if isinstance(value, str) and value.startswith("="):
                            value = values.cell(row, columns[key]).value
                        record[key] = value
                    record["_row"] = snapshot_detail_row(source, values, row, last_column)
                    consolidated[row_day].append(record)
        finally:
            formulas.close()
            cached.close()

    if not sources and not consolidated:
        raise ReportError(f"所选日报中没有 {year}/{month} 的数据。")

    for day in sorted(set(sources) | set(consolidated)):
        if day not in sources:
            logger(f"  正在读取日报内部日期：{day.month}.{day.day}")
            rows.extend(consolidated[day])
            continue
        daily_path, name = sources[day]
        logger(f"  正在读取：{daily_path.name} / {name}")
        formulas = load_workbook(daily_path, data_only=False)
        cached = load_workbook(daily_path, data_only=True)
        try:
            source, values = formulas[name], cached[name]
            header_row, columns = find_header(source, DAILY_ALIASES)
            clean_key = (str(daily_path.resolve()).casefold(), name.casefold())
            _clean_second_header(
                source, values, header_row, logger, clean_key not in cleaning_reported
            )
            cleaning_reported.add(clean_key)
            header_row, columns = find_header(source, DAILY_ALIASES)
            date_column = find_optional_header_column(source, header_row, DATED_DAILY_ALIASES["date"])
            if date_column is not None:
                columns["date"] = date_column
            summary_row = find_summary_row(source, header_row) or (source.max_row + 1)
            refund_found = refund_found or "refund" in columns
            last_column = last_business_column(source, header_row)
            if data_columns is None:
                data_columns = dict(columns)
                column_layout = source_column_layout(source, last_column)
            for row in range(header_row + 1, summary_row):
                store = source.cell(row, columns["store"]).value
                if store is None or str(store).strip() == "" or is_summary_label(store):
                    continue
                record = {"store": str(store).strip()}
                for key in ("sales", "orders", "refund"):
                    if key not in columns:
                        continue
                    value = source.cell(row, columns[key]).value
                    if isinstance(value, str) and value.startswith("="):
                        value = values.cell(row, columns[key]).value
                    record[key] = value
                record["_row"] = snapshot_detail_row(source, values, row, last_column)
                rows.append(record)
        finally:
            formulas.close()
            cached.close()
    if data_columns is None or column_layout is None:
        raise ReportError("日报中没有可复制的明细数据结构。")
    if "date" not in data_columns:
        raise ReportError("日报中找不到日期列，无法写入所选月报日期。")
    return rows, refund_found, data_columns, column_layout


def _create_month_data_sheet(workbook, name: str):
    if name in workbook.sheetnames:
        del workbook[name]
    return workbook.create_sheet(name)


def _write_month_data(sheet, rows: list[dict], column_layout: dict[int, dict],
                      date_column: int, date_label: str):
    write_complete_detail_rows(
        sheet, [record["_row"] for record in rows], column_layout, date_column, date_label
    )


def _monthly_numeric(value) -> float:
    if value is None or isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").replace("￥", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _month_row_snapshot(sheet, row: int) -> dict:
    cells = []
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        cells.append({
            "value": cell.value, "coordinate": cell.coordinate,
            "font": copy(cell.font), "fill": copy(cell.fill),
            "border": copy(cell.border), "alignment": copy(cell.alignment),
            "number_format": cell.number_format, "protection": copy(cell.protection),
            "comment": copy(cell.comment), "hyperlink": copy(cell.hyperlink),
        })
    dimension = sheet.row_dimensions[row]
    return {
        "cells": cells, "height": dimension.height, "hidden": dimension.hidden,
        "outline_level": dimension.outlineLevel, "collapsed": dimension.collapsed,
    }


def _restore_month_row(sheet, row: int, snapshot: dict) -> None:
    for column, source in enumerate(snapshot["cells"], 1):
        cell = sheet.cell(row, column)
        value = source["value"]
        if isinstance(value, str) and value.startswith("="):
            try:
                value = Translator(value, origin=source["coordinate"]).translate_formula(cell.coordinate)
            except (TypeError, ValueError):
                pass
        cell.value = value
        cell.font = copy(source["font"]); cell.fill = copy(source["fill"])
        cell.border = copy(source["border"]); cell.alignment = copy(source["alignment"])
        cell.number_format = source["number_format"]
        cell.protection = copy(source["protection"])
        cell.comment = copy(source["comment"]); cell._hyperlink = copy(source["hyperlink"])
    dimension = sheet.row_dimensions[row]
    dimension.height = snapshot["height"]; dimension.hidden = snapshot["hidden"]
    dimension.outlineLevel = snapshot["outline_level"]; dimension.collapsed = snapshot["collapsed"]


def _sort_and_date_month(sheet, rows: list[dict], year: int, month: int,
                         logger: Callable[[str], None]):
    header_row, columns = _find_month_report_header(sheet)
    summary_row = find_summary_row(sheet, header_row)
    if summary_row is None:
        last_store = max(
            (row for row in range(header_row + 1, sheet.max_row + 1)
             if sheet.cell(row, columns["store"]).value not in (None, "")),
            default=header_row,
        )
        summary_row = last_store + 1
    start_row, end_row = header_row + 1, summary_row - 1
    totals: dict[str, float] = defaultdict(float)
    for record in rows:
        totals[record["store"]] += _monthly_numeric(record.get("sales"))
    stores, blanks = [], []
    for row in range(start_row, summary_row):
        store = sheet.cell(row, columns["store"]).value
        item = (str(store).strip() if store is not None else "", _month_row_snapshot(sheet, row))
        (stores if item[0] and not is_summary_label(item[0]) else blanks).append(item)
    stores.sort(key=lambda item: totals.get(item[0], 0.0), reverse=True)
    for row, (_, snapshot) in zip(range(start_row, summary_row), stores + blanks):
        _restore_month_row(sheet, row, snapshot)

    if columns["store"] >= 3:
        date_text = f"{year}年{month}月"
        rank = 1
        for row in range(start_row, summary_row):
            sheet.cell(row, 1).value = date_text
            store = sheet.cell(row, columns["store"]).value
            if store not in (None, "") and not is_summary_label(store):
                sheet.cell(row, 2).value = rank
                rank += 1
    logger("排序字段：\n本月销售额")
    logger(f"排序范围：\n第{start_row}行 至 第{end_row}行")
    logger(f"汇总行：\n第{summary_row}行")
    logger(f"排名填充：\nB{start_row}:B{end_row}")
    return header_row, summary_row, columns, [totals.get(name, 0.0) for name, _ in stores]


def _update_month_formulas(sheet, previous_name: str, data_name: str, include_refund: bool,
                           data_columns: dict[str, int]):
    header_row, columns = _find_month_report_header(sheet)
    summary_row = find_summary_row(sheet, header_row) or (sheet.max_row + 1)
    store_letter = get_column_letter(data_columns["store"])
    sales_letter = get_column_letter(columns["sales"])
    previous_letter = get_column_letter(columns["previous"])
    for row in range(header_row + 1, summary_row):
        store = sheet.cell(row, columns["store"]).value
        if store is None or str(store).strip() == "":
            continue
        report_store_letter = get_column_letter(columns["store"])
        store_ref = f"${report_store_letter}{row}"
        previous = sheet.cell(row, columns["previous"])
        if isinstance(previous.value, str) and previous.value.startswith("="):
            previous.value = replace_sheet_references(previous.value, previous_name)
        targets = (("sales", data_columns["sales"]), ("orders", data_columns["orders"]))
        if include_refund and "refund" in columns:
            targets += (("refund", data_columns["refund"]),)
        for key, data_column in targets:
            cell = sheet.cell(row, columns[key])
            value_letter = get_column_letter(data_column)
            cell.value = (f"=SUMIF('{data_name}'!${store_letter}:${store_letter},{store_ref},"
                          f"'{data_name}'!${value_letter}:${value_letter})")

        if include_refund and "refund" in columns and "refund_rate" in columns:
            refund_letter = get_column_letter(columns["refund"])
            sheet.cell(row, columns["refund_rate"]).value = (
                f'=IFERROR(({refund_letter}{row}/{sales_letter}{row}),"0")'
            )
        if "comparison" in columns:
            sheet.cell(row, columns["comparison"]).value = (
                f'=IFERROR(({sales_letter}{row}-{previous_letter}{row})/'
                f'{previous_letter}{row},"0")'
            )

    start_row, end_row = header_row + 1, summary_row - 1
    for key in ("sales", "orders", "previous"):
        column_letter = get_column_letter(columns[key])
        sheet.cell(summary_row, columns[key]).value = (
            f"=SUM({column_letter}{start_row}:{column_letter}{end_row})"
        )
    if include_refund and "refund" in columns:
        refund_letter = get_column_letter(columns["refund"])
        sheet.cell(summary_row, columns["refund"]).value = (
            f"=SUM({refund_letter}{start_row}:{refund_letter}{end_row})"
        )
        if "refund_rate" in columns:
            sheet.cell(summary_row, columns["refund_rate"]).value = (
                f'=IFERROR(({refund_letter}{summary_row}/{sales_letter}{summary_row}),"0")'
            )
    if "comparison" in columns:
        sheet.cell(summary_row, columns["comparison"]).value = (
            f'=IFERROR(({sales_letter}{summary_row}-{previous_letter}{summary_row})/'
            f'{previous_letter}{summary_row},"0")'
        )
    return header_row, summary_row, columns


def process_monthly(daily_paths: list[Path] | Path, report_path: Path, year: int, month: int,
                    output_path: Path, logger: Callable[[str], None] = lambda _: None) -> Path:
    if isinstance(daily_paths, Path):
        daily_paths = [daily_paths]
    target_name, data_name = f"{month}月", str(month)
    logger("[1/10]\n正在读取月报模板...")
    logger(f"正在打开文件：\n{report_path.name}")
    sheet_names = xlsx_sheet_names(report_path)
    candidates = [name for name in sheet_names
                  if MONTH_PATTERN.fullmatch(name) and name != target_name]
    if not candidates:
        raise ReportError("月报文件中找不到上月模板工作表。")
    previous_name = candidates[-1]
    logger(f"发现工作表：\n{previous_name}")
    logger(f"正在读取：\n{previous_name}")
    workbook = load_template_workbook(report_path, logger)
    logger("读取完成。")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=output_path.parent, suffix=".xlsx", delete=False) as handle:
        working_path = Path(handle.name)
    try:
        previous = workbook[previous_name]
        logger(f"[2/10]\n正在复制工作表{previous.title}...")
        target = duplicate_template_sheet(workbook, previous, target_name)
        cached_workbook = load_workbook(report_path, data_only=True, keep_links=False)
        try:
            solidify_formula_sheet(previous, cached_workbook[previous_name])
        finally:
            cached_workbook.close()
        keep_only_sheets(workbook, {previous_name, target_name})
        logger(f"[3/10]\n正在创建{data_name}...")
        data_sheet = _create_month_data_sheet(workbook, data_name)
        logger("[4/10]\n正在读取日报...")
        rows, daily_has_refund, data_columns, column_layout = _collect_month(
            daily_paths, year, month, logger
        )
        logger("[5/10]\n正在删除汇总行...")
        _, report_columns = _find_month_report_header(target)
        field_log = [
            f"店铺={get_column_letter(report_columns['store'])}列",
            f"销售额={get_column_letter(report_columns['sales'])}列",
            f"订单={get_column_letter(report_columns['orders'])}列",
            f"上月销售={get_column_letter(report_columns['previous'])}列",
        ]
        for key, label in (("refund", "退款金额"), ("refund_rate", "退款率"),
                           ("comparison", "环比")):
            if key in report_columns:
                field_log.append(f"{label}={get_column_letter(report_columns[key])}列")
        logger("月报字段识别：\n" + "\n".join(field_log))
        include_refund = daily_has_refund and "refund" in report_columns
        if not include_refund:
            logger("未检测到可同时使用的退款字段，跳过退款公式。")
        logger("[6/10]\n正在复制数据...")
        _write_month_data(
            data_sheet, rows, column_layout, data_columns["date"], target_name
        )
        sort_header, sort_summary, sort_columns, ordered_totals = _sort_and_date_month(
            target, rows, year, month, logger
        )
        logger("[7/10]\n正在修改SUMIF公式...")
        header_row, summary_row, columns = _update_month_formulas(
            target, previous.title, data_name, include_refund, data_columns
        )
        logger("[8/10]\n正在修改IFERROR公式...")
        selected_columns = [columns["previous"], columns["sales"], columns["orders"]]
        if include_refund and "refund" in columns:
            selected_columns.append(columns["refund"])
            if "refund_rate" in columns:
                selected_columns.append(columns["refund_rate"])
        if "comparison" in columns:
            selected_columns.append(columns["comparison"])
        normalize_division_columns(target, selected_columns, header_row + 1, summary_row)
        if len(list(target.conditional_formatting)) < len(list(previous.conditional_formatting)):
            raise ReportError("模板条件格式未完整保留。")
        if len(target.data_validations.dataValidation) < len(previous.data_validations.dataValidation):
            raise ReportError("模板数据验证未完整保留。")
        formulas = [target.cell(row, column).value
                    for column in selected_columns
                    for row in range(header_row + 1, summary_row)
                    if isinstance(target.cell(row, column).value, str)
                    and target.cell(row, column).value.startswith("=")]
        if not any("SUMIF(" in value.upper() and f"'{data_name}'!" in value for value in formulas):
            raise ReportError("本月 SUMIF 数据引用检查失败。")
        if not any("VLOOKUP(" in value.upper() and f"'{previous.title}'!" in value for value in formulas):
            raise ReportError("上月 VLOOKUP 引用检查失败。")
        if output_path.name != report_path.name:
            raise ReportError("输出文件名发生变化。")
        if set(workbook.sheetnames) != {previous_name, target_name, data_name}:
            raise ReportError("月报输出工作表数量或名称不正确。")
        required_daily_fields = {
            "date": "日期", "store": "店铺名称", "sales": "销售额", "orders": "订单数",
        }
        missing_daily_fields = [
            label for key, label in required_daily_fields.items()
            if key not in data_columns or not isinstance(data_columns[key], int)
            or data_columns[key] < 1
        ]
        if missing_daily_fields:
            raise ReportError("月报日报数据缺少必要字段：" + "、".join(missing_daily_fields))
        if any(data_sheet.cell(row, data_columns["date"]).value != target_name
               for row in range(1, data_sheet.max_row + 1)):
            raise ReportError("月报数据工作表日期列检查失败。")
        if ordered_totals != sorted(ordered_totals, reverse=True):
            raise ReportError("月报本月销售额降序检查失败。")
        if sort_columns["store"] >= 3:
            expected_date = f"{year}年{month}月"
            if any(target.cell(row, 1).value != expected_date
                   for row in range(sort_header + 1, sort_summary)):
                raise ReportError("月报第一列年月检查失败。")
            ranks = [target.cell(row, 2).value for row in range(sort_header + 1, sort_summary)]
            if ranks != list(range(1, len(ranks) + 1)):
                raise ReportError("月报排名连续性检查失败。")
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        logger("[9/10]\n正在保存Excel...")
        workbook.save(working_path)
    finally:
        workbook.close()
    working_path.replace(output_path)
    if not output_path.is_file() or output_path.name != report_path.name:
        raise ReportError("月报未按原文件名成功输出。")
    logger(
        "月报生成成功\n识别字段：\n"
        f"店铺名称={get_column_letter(columns['store'])}列\n"
        f"销售额={get_column_letter(columns['sales'])}列\n"
        f"订单数={get_column_letter(columns['orders'])}列\n"
        f"上月销售={get_column_letter(columns['previous'])}列"
    )
    logger("[10/10]\n完成。")
    return output_path
