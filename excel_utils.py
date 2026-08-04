"""无业务含义的 Excel 格式、日期和路径工具。"""
from __future__ import annotations

from copy import copy
from datetime import date, datetime
import os
from pathlib import Path
import re
from threading import Event, Timer
from typing import Callable
import winreg
import xml.etree.ElementTree as ET
from zipfile import BadZipFile, ZipFile
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def xlsx_sheet_names(filename: Path) -> list[str]:
    """Quickly read worksheet names without parsing worksheet XML content."""
    try:
        with ZipFile(filename) as archive:
            root = ET.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, ET.ParseError, OSError) as exc:
        raise ReportError(f"无法读取Excel工作表列表：{filename.name}") from exc
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [node.attrib["name"] for node in root.findall("x:sheets/x:sheet", namespace)]


def load_template_workbook(filename: Path, logger: Callable[[str], None], slow_seconds: float = 60):
    """Load an editable template once and report unusually slow loading."""
    completed = Event()

    def warn_if_slow() -> None:
        if not completed.is_set():
            logger("Excel模板较大，正在处理。")

    timer = Timer(slow_seconds, warn_if_slow)
    timer.daemon = True
    timer.start()
    try:
        return load_workbook(filename, data_only=False, keep_links=False)
    finally:
        completed.set()
        timer.cancel()


class ReportError(RuntimeError):
    """可直接显示给用户的输入或模板错误。"""


def desktop_path() -> Path:
    with winreg.OpenKey(
        winreg.HKEY_CURRENT_USER,
        r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
    ) as key:
        value, _ = winreg.QueryValueEx(key, "Desktop")
    return Path(os.path.expandvars(value)).expanduser()


def daily_sheet_name(day: date) -> str:
    return f"{day.month}.{day.day}"


def normalized(value) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def parse_date_value(value, default_year: int) -> date | None:
    """解析日报内部日期字段，兼容日期对象、Excel序列值和常见文本格式。"""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        integer = int(value)
        if 19000101 <= integer <= 29991231:
            try:
                return date(integer // 10000, integer // 100 % 100, integer % 100)
            except ValueError:
                return None
        if 1 <= value <= 100000:
            try:
                parsed = from_excel(value)
                return parsed.date() if isinstance(parsed, datetime) else parsed
            except (ValueError, OverflowError):
                return None
    text = str(value or "").strip()
    for pattern in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})", text)
    if not match:
        match = re.fullmatch(r"(\d{1,2})月(\d{1,2})日?", text)
    if match:
        try:
            return date(default_year, int(match.group(1)), int(match.group(2)))
        except ValueError:
            return None
    return None


def is_summary_label(value) -> bool:
    text = str(value or "").strip()
    return any(word in text for word in ("合计", "汇总", "总计"))


def find_header(sheet, aliases: dict[str, tuple[str, ...]]) -> tuple[int, dict[str, int]]:
    """Scan the first 100 rows and prefer earlier (more specific) aliases."""
    required = {key for key in aliases if key != "refund"}
    for row in range(1, min(sheet.max_row, 100) + 1):
        found: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                expected = normalized(name).casefold()
                matching = [column for column in range(1, sheet.max_column + 1)
                            if normalized(sheet.cell(row, column).value).casefold() == expected]
                if matching:
                    found[key] = matching[0]
                    break
        if required.issubset(found):
            return row, found
    missing = "、".join(sorted(required))
    preview_lines = []
    for row in range(1, min(sheet.max_row, 20) + 1):
        values = [str(sheet.cell(row, column).value or "")
                  for column in range(1, min(sheet.max_column, 30) + 1)]
        preview_lines.append(f"第{row}行：" + " | ".join(values))
    preview = "\n".join(preview_lines) or "（工作表为空）"
    raise ReportError(
        f"工作表“{sheet.title}”前100行找不到必要字段：{missing}\n"
        f"实际读取到的前20行内容：\n{preview}"
    )


def find_optional_header_column(sheet, header_row: int, names: tuple[str, ...]) -> int | None:
    expected = {normalized(name) for name in names}
    for column in range(1, sheet.max_column + 1):
        if normalized(sheet.cell(header_row, column).value) in expected:
            return column
    return None


def solidify_formula_sheet(target, cached) -> int:
    """Replace formulas in a historical sheet with the source file's displayed cache values."""
    changed = 0
    for cell in tuple(target._cells.values()):
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = cached.cell(cell.row, cell.column).value
            changed += 1
    return changed


def keep_only_sheets(workbook, names: set[str]) -> None:
    """Delete unneeded sheets only after the template copy and history solidification."""
    for sheet in tuple(workbook.worksheets):
        if sheet.title not in names:
            workbook.remove(sheet)


def find_summary_row(sheet, header_row: int) -> int | None:
    result = None
    for row in range(header_row + 1, sheet.max_row + 1):
        text = "".join(
            str(value or "")
            for value in (sheet.cell(row, column).value for column in range(1, sheet.max_column + 1))
            if not (isinstance(value, str) and value.startswith("="))
        )
        if any(word in text for word in ("合计", "汇总", "总计")):
            result = row
    return result


def duplicate_template_sheet(workbook, source, target_name: str):
    """严格用 copy_worksheet 复制模板，并补齐其默认遗漏的对象。"""
    source_index = workbook.worksheets.index(source)
    duplicate = workbook.copy_worksheet(source)
    duplicate.conditional_formatting = copy(source.conditional_formatting)
    duplicate.data_validations = copy(source.data_validations)
    duplicate.freeze_panes = source.freeze_panes
    duplicate.sheet_format = copy(source.sheet_format)
    duplicate.sheet_properties = copy(source.sheet_properties)
    duplicate.views = copy(source.views)
    duplicate.page_margins = copy(source.page_margins)
    duplicate.page_setup = copy(source.page_setup)
    duplicate.print_options = copy(source.print_options)
    duplicate.auto_filter = copy(source.auto_filter)
    duplicate.protection = copy(source.protection)
    temporary = f"{source.title} (2)"
    counter = 2
    while temporary in workbook.sheetnames and workbook[temporary] is not duplicate:
        counter += 1
        temporary = f"{source.title} ({counter})"
    duplicate.title = temporary
    if target_name in workbook.sheetnames and workbook[target_name] is not duplicate:
        del workbook[target_name]
    duplicate.title = target_name
    workbook._sheets.remove(duplicate)
    workbook._sheets.insert(source_index + 1, duplicate)
    return duplicate


def copy_cell_format(source, target) -> None:
    if isinstance(source, MergedCell):
        return
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def last_business_column(sheet, header_row: int) -> int:
    """Return the last column containing a header or a value below the header."""
    last = 1
    for row in range(header_row, sheet.max_row + 1):
        for column in range(sheet.max_column, last, -1):
            if sheet.cell(row, column).value is not None:
                last = column
                break
    return last


def snapshot_detail_row(source, cached, row: int, last_column: int) -> dict:
    """Capture a complete detail row, including formatting and applicable CF rules."""
    cells = []
    for column in range(1, last_column + 1):
        original = source.cell(row, column)
        value = original.value
        if isinstance(value, str) and value.startswith("="):
            value = cached.cell(row, column).value
        cells.append({
            "value": value,
            "font": copy(original.font),
            "fill": copy(original.fill),
            "border": copy(original.border),
            "alignment": copy(original.alignment),
            "number_format": original.number_format,
            "protection": copy(original.protection),
            "comment": copy(original.comment),
            "hyperlink": copy(original.hyperlink),
        })
    conditional_rules = []
    for conditional in source.conditional_formatting:
        for cell_range in conditional.sqref.ranges:
            if cell_range.min_row <= row <= cell_range.max_row:
                min_column = max(1, cell_range.min_col)
                max_column = min(last_column, cell_range.max_col)
                if min_column <= max_column:
                    for rule in source.conditional_formatting[conditional]:
                        conditional_rules.append((min_column, max_column, copy(rule)))
    dimension = source.row_dimensions[row]
    return {
        "cells": cells,
        "height": dimension.height,
        "hidden": dimension.hidden,
        "outline_level": dimension.outlineLevel,
        "collapsed": dimension.collapsed,
        "conditional_rules": conditional_rules,
    }


def write_complete_detail_rows(sheet, rows: list[dict], column_layout: dict[int, dict],
                               date_column: int, date_label: str) -> None:
    """Append full source rows without creating or retaining a header row."""
    for column, layout in column_layout.items():
        letter = get_column_letter(column)
        target_dimension = sheet.column_dimensions[letter]
        target_dimension.width = layout.get("width")
        target_dimension.hidden = layout.get("hidden", False)
        target_dimension.outlineLevel = layout.get("outline_level", 0)
        target_dimension.collapsed = layout.get("collapsed", False)
    for target_row, record in enumerate(rows, 1):
        for column, snapshot in enumerate(record["cells"], 1):
            target = sheet.cell(target_row, column)
            if snapshot["value"] is not None:
                target.value = snapshot["value"]
            target.font = copy(snapshot["font"])
            target.fill = copy(snapshot["fill"])
            target.border = copy(snapshot["border"])
            target.alignment = copy(snapshot["alignment"])
            target.number_format = snapshot["number_format"]
            target.protection = copy(snapshot["protection"])
            target.comment = copy(snapshot["comment"])
            target._hyperlink = copy(snapshot["hyperlink"])
        sheet.cell(target_row, date_column).value = date_label
        dimension = sheet.row_dimensions[target_row]
        dimension.height = record["height"]
        dimension.hidden = record["hidden"]
        dimension.outlineLevel = record["outline_level"]
        dimension.collapsed = record["collapsed"]
        for min_column, max_column, rule in record["conditional_rules"]:
            start = f"{get_column_letter(min_column)}{target_row}"
            end = f"{get_column_letter(max_column)}{target_row}"
            sheet.conditional_formatting.add(f"{start}:{end}", copy(rule))


def source_column_layout(sheet, last_column: int) -> dict[int, dict]:
    result = {}
    for column in range(1, last_column + 1):
        dimension = sheet.column_dimensions[get_column_letter(column)]
        result[column] = {
            "width": dimension.width,
            "hidden": dimension.hidden,
            "outline_level": dimension.outlineLevel,
            "collapsed": dimension.collapsed,
        }
    return result


def choose_latest_sheet(workbook, pattern: re.Pattern[str], excluded: set[str] | None = None):
    excluded = excluded or set()
    candidates = [sheet for sheet in workbook.worksheets
                  if sheet.title not in excluded and pattern.fullmatch(sheet.title)]
    if not candidates:
        raise ReportError("找不到可复制的上一期模板工作表。")
    return candidates[-1]


def brand_key(path: Path, kind: str) -> str:
    stem = path.stem.strip()
    for suffix in (f"品牌{kind}", kind, "品牌日报", "日报"):
        if stem.endswith(suffix):
            stem = stem[:-len(suffix)]
            break
    return stem.strip()


def pair_files(daily_files: list[Path], report_files: list[Path], kind: str) -> list[tuple[Path, Path]]:
    daily_map = {brand_key(path, "日报"): path for path in daily_files}
    pairs = []
    missing = []
    for report in report_files:
        key = brand_key(report, kind)
        daily = daily_map.get(key)
        if daily is None:
            missing.append(report.name)
        else:
            pairs.append((daily, report))
    if missing:
        raise ReportError("以下报表找不到同品牌日报：" + "、".join(missing))
    return pairs
