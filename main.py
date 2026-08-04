"""Excel月报自动汇总助手 GUI。"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import queue
import re
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from excel_utils import ReportError, desktop_path
from file_match import match_brand_files, report_keyword
from monthly_report import process_monthly


def _packaged_excel_self_test(output_dir: Path) -> None:
    """Verify that the packaged executable can create, read, and save XLSX files."""
    from openpyxl import Workbook, load_workbook

    output_dir.mkdir(parents=True, exist_ok=True)
    source = output_dir / "exe_openpyxl_read_test.xlsx"
    saved = output_dir / "exe_openpyxl_save_test.xlsx"
    result = output_dir / "exe_openpyxl_self_test.txt"
    workbook = Workbook()
    workbook.active["A1"] = "openpyxl packaged"
    workbook.active["B1"] = 123
    workbook.save(source)
    workbook.close()
    loaded = load_workbook(source, data_only=False)
    if loaded.active["A1"].value != "openpyxl packaged" or loaded.active["B1"].value != 123:
        raise RuntimeError("The packaged executable failed to read XLSX values.")
    loaded.active["A2"] = "save verified"
    loaded.save(saved)
    loaded.close()
    verified = load_workbook(saved, data_only=False, read_only=True)
    if verified.active["A2"].value != "save verified":
        verified.close()
        raise RuntimeError("The packaged executable failed to save XLSX values.")
    verified.close()
    result.write_text("OPENPYXL_EXE_SELF_TEST_OK", encoding="utf-8")


class MonthlyReportFrame(ttk.Frame):
    def __init__(self, master, event_queue):
        super().__init__(master, padding=16)
        self.events = event_queue
        self.daily_files: list[Path] = []
        self.report_files: list[Path] = []
        self._build()

    def _build(self):
        ttk.Label(self, text="月报汇总功能",
                  font=("Microsoft YaHei UI", 16, "bold")).pack(anchor="w")
        date_frame = ttk.LabelFrame(self, text="处理月份", padding=10)
        date_frame.pack(fill="x", pady=(12, 8))
        ttk.Label(date_frame, text="目标月份").grid(row=0, column=0, sticky="w")
        self.month_entry = ttk.Entry(date_frame, width=16)
        self.month_entry.insert(0, date.today().strftime("%Y/%m"))
        self.month_entry.grid(row=0, column=1, padx=8)

        files = ttk.LabelFrame(self, text="文件选择", padding=10)
        files.pack(fill="x")
        ttk.Button(files, text="1. 选择多个品牌日报", command=self.select_daily).grid(
            row=0, column=0, sticky="ew", pady=4
        )
        self.daily_label = ttk.Label(files, text="尚未选择")
        self.daily_label.grid(row=0, column=1, sticky="w", padx=10)
        ttk.Button(files, text="2. 选择多个品牌月报", command=self.select_reports).grid(
            row=1, column=0, sticky="ew", pady=4
        )
        self.report_label = ttk.Label(files, text="尚未选择")
        self.report_label.grid(row=1, column=1, sticky="w", padx=10)
        files.columnconfigure(1, weight=1)

        self.start_button = ttk.Button(self, text="开始月报汇总", command=self.start)
        self.start_button.pack(anchor="w", pady=(12, 6))
        self.progress = ttk.Progressbar(self, mode="determinate", maximum=10)
        self.progress.pack(fill="x")
        ttk.Label(self, text="输出位置：Windows桌面；输出文件名保持原月报名称").pack(
            anchor="w", pady=(6, 8)
        )
        self.log = tk.Text(self, height=14, state="disabled", wrap="word")
        self.log.pack(fill="both", expand=True)

    def select_daily(self):
        paths = filedialog.askopenfilenames(parent=self, title="选择品牌日报",
                                            filetypes=[("Excel 工作簿", "*.xlsx")])
        if paths:
            self.daily_files = [Path(path) for path in paths]
            self.daily_label.config(text=f"已选择 {len(paths)} 个文件")

    def select_reports(self):
        paths = filedialog.askopenfilenames(parent=self, title="选择品牌月报",
                                            filetypes=[("Excel 工作簿", "*.xlsx")])
        if paths:
            self.report_files = [Path(path) for path in paths]
            self.report_label.config(text=f"已选择 {len(paths)} 个文件")

    def start(self):
        if not self.daily_files or not self.report_files:
            messagebox.showwarning("缺少文件", "请选择品牌日报和品牌月报。", parent=self)
            return
        try:
            self.append_log("正在匹配品牌...")
            self.pairs = match_brand_files(self.daily_files, self.report_files)
            for matches, report in self.pairs:
                self.append_log(f"月报文件：{report.name}")
                self.append_log(f"识别品牌：{report_keyword(report)}")
                self.append_log("匹配日报：")
                for matched in matches:
                    self.append_log(f"  {matched.name}")
                self.append_log("")
            parsed = datetime.strptime(self.month_entry.get().strip(), "%Y/%m")
            self.year, self.month = parsed.year, parsed.month
        except ValueError:
            messagebox.showerror("日期错误", "请输入正确月份：YYYY/MM。", parent=self)
            return
        except ReportError as exc:
            messagebox.showerror("文件匹配失败", str(exc), parent=self)
            return
        self.start_button.config(state="disabled")
        self.progress["value"] = 0
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            desktop = desktop_path()
            outputs = []
            for daily_group, report in self.pairs:
                output = desktop / report.name
                logger = lambda text: self.events.put((self, "log", text))
                outputs.append(process_monthly(
                    daily_group, report, self.year, self.month, output, logger
                ))
            self.events.put((self, "done", f"处理完成，共输出 {len(outputs)} 个文件到桌面。"))
        except Exception as exc:
            self.events.put((self, "error", str(exc)))

    def finish(self):
        self.start_button.config(state="normal")

    def append_log(self, text):
        match = re.match(r"\[(\d+)/10\]", text)
        if match:
            self.progress["value"] = int(match.group(1))
        self.log.config(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.config(state="disabled")


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel月报自动汇总助手")
        self.geometry("820x610")
        self.minsize(720, 540)
        self.events = queue.Queue()
        MonthlyReportFrame(self, self.events).pack(fill="both", expand=True)
        self.after(100, self._drain_events)

    def _drain_events(self):
        try:
            while True:
                frame, kind, text = self.events.get_nowait()
                frame.append_log(text)
                if kind in ("done", "error"):
                    frame.finish()
                    if kind == "done":
                        messagebox.showinfo("完成", text, parent=frame)
                    else:
                        messagebox.showerror("处理失败", text, parent=frame)
        except queue.Empty:
            pass
        self.after(100, self._drain_events)


if __name__ == "__main__":
    if len(sys.argv) == 3 and sys.argv[1] == "--self-test-excel":
        _packaged_excel_self_test(Path(sys.argv[2]))
    else:
        Application().mainloop()
